import requests
import re
from bs4 import BeautifulSoup
import openpyxl
import datetime
import re

now = datetime.datetime.now()
date = now.strftime('%Y-%m-%d %H-%M-%S')

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '상한가'

row = 1
sheet.cell(row, 1, '종목 코드')
sheet.cell(row, 2, '종목 이름')
sheet.cell(row, 3, '한 주당 가격')
sheet.cell(row, 4, '전날 대비 가격 변동')
sheet.cell(row, 5, '전날 대비 등락')
sheet.cell(row, 6, '관련 된 뉴스기사')
sheet.cell(row, 7, '일봉 차트')
sheet.cell(row, 8, '주봉 차트')
sheet.cell(row, 9, '월봉 차트')

def get_bs_obj(): 

    url = "https://finance.naver.com/sise/sise_upper.nhn"
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")
    units_up = bs_obj.find_all('table', {'class':'type_5'})
    
    for unit in units_up:
        tr_up = unit.find_all('tr')
        #header = units_up.select('th')
        print('tr_up: %s' %tr_up)


 

units_up = get_bs_obj()
