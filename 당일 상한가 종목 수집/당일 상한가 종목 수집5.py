import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
import re

now = datetime.datetime.now()
date = now.strftime('%Y-%m-%d %H-%M-%S')

URL = 'https://finance.naver.com/'

raw = requests.get(URL, verify=False)

html = BeautifulSoup(raw.text,'lxml')

units_up = html.select('#_topItems2>tr')
#print('units_up : %s' % units_up)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '상한가'

'''
상한가를 나타내는 아이콘이 있는지 없는지를 확인 True/False를 반환하여 개수를 확인한다.
'''
check_str = html.find_all('em', attrs={'class':'bu_p bu_pup2'})
print('check_str:%s'%len(check_str))

row = 1
sheet.cell(row, 1, '종목 코드')
sheet.cell(row, 2, '종목 이름')
sheet.cell(row, 3, '한 주당 가격')
sheet.cell(row, 4, '전날 대비 가격 변동')
sheet.cell(row, 5, '전날 대비 등락')
sheet.cell(row, 6, '관련 된 뉴스기사')
sheet.cell(row, 7, '일봉 차트')
sheet.cell(row, 8, '주봉 차트')
sheet.cell(row, 9, '월봉 차트')
row += 1

for unit in units_up[:len(check_str)]:
    title_up = unit.select_one('#_topItems2 > tr> th > a').text
    price_up = unit.select_one('#_topItems2 > tr> td')
    up = unit.select_one('#_topItems2 > tr > td:nth-child(3)').text
    percent_up = unit.select_one('#_topItems2 > tr> td:nth-child(4)')
    code_up = unit.select_one('#_topItems2 > tr > th > a')
    code_up_href = code_up['href']
    #print('code_up_href: %s' %code_up_href)
    code6 = code_up_href[-6:]

    up = up.replace('상한가', '↑')

    news_up = 'https://search.naver.com/search.naver?sm=tab_hty.top&where=news&query=' + title_up
    raw2 = requests.get(news_up)
    html2= BeautifulSoup(raw2.text,'lxml')
    news_up_box = html2.find('div',{'class':'group_news'})
    news_up_list = news_up_box.find_all('div',{'class':'news_area'})

    sheet.cell(row, 1, code6)
    sheet.cell(row, 2, title_up)
    sheet.cell(row, 3, price_up.text + '원')
    sheet.cell(row, 4, up)
    sheet.cell(row, 5, percent_up.text)

    news_links = []
    for new in news_up_list[:3]:
        new_title_up = new.find('a',{'class' : 'news_tit'})
        #print("new_title_up.text:%s" % new_title_up.text)
        link_up = new.find('a',{'class' : 'api_txt_lines dsc_txt_wrap'})
        news_links.append(link_up['href'])
        sheet.cell(row, 6, '=HYPERLINK("{}","{}")'.format(link_up['href'], new_title_up.text))

    title_up = unit.select_one('#_topItems2 > tr> th > a')
    chart_up_url = 'https://finance.naver.com' + title_up['href']
    chart_up_raw = requests.get(chart_up_url)
    chart_up_html = BeautifulSoup(chart_up_raw.text,'lxml')
    chart_up = chart_up_html.select_one('#img_chart_area')
    chart_up = chart_up['src']
    chart_up_day = chart_up.replace('area','candle')
    chart_up_week = chart_up_day.replace('day','week')
    chart_up_month = chart_up_day.replace('day','month')

    sheet.cell(row, 7, '=HYPERLINK("{}","일봉 차트")'.format(chart_up_day))
    sheet.cell(row, 8, '=HYPERLINK("{}","주봉 차트")'.format(chart_up_week))
    sheet.cell(row, 9, '=HYPERLINK("{}","월봉 차트")'.format(chart_up_month))
    row += 1

# 현재 시간을 가져온 뒤, 시간 형식을 변경합니다.
now = datetime.datetime.now()
date_time = now.strftime("%Y%m%d_%H%M%S")

wb.save(f'D:/lastBox/당일 상한가 종목 수집/당일 상한가 종목 수집 xlsx/상한가_{date_time}.xlsx')

