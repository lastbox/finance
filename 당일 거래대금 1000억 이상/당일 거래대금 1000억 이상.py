import requests
import re
import openpyxl
import datetime
from bs4 import BeautifulSoup

now = datetime.datetime.now()
date_time = now.strftime('%Y-%m-%d %H-%M-$S')

#URL = 'https://finance.naver.com/sise/sise_rise.naver'
URL = 'https://www.kokstock.com/stock/amount.asp?search_base=0&page=1&pagesize=20&view_quant=0&search_tp=&search_quant=6&search_sort=5'

raw = requests.get(URL, verify=False)
html = BeautifulSoup(raw.text, 'lxml')

wb = openpyxl.Workbook()
sheet = wb.active

# #contentarea > div.box_type_l > table > tbody > tr:nth-child(1) # 첫번째 tr임...
# 후에 순번으로 찾을때 필요할 수 있어 주석으로 남겨둠
top_money_table = html.select('#divList > div.table-responsive.board-table > table > tbody > tr')
#top_money_tr = top_money_table.select('tr')
#print('top_money_table:%s'%top_money_table)


row = 1
sheet.cell(row, 1, 'Market')
sheet.cell(row, 2, '종목 코드')
sheet.cell(row, 3, '종목 이름')
sheet.cell(row, 4, '현재 가격')
sheet.cell(row, 5, '등락률')
sheet.cell(row, 6, '거래대금 (억)')
sheet.cell(row, 7, '시가총액 (억)')
row += 1

for unit in top_money_table[:len(top_money_table)]:
    pass
    #print("dk")
    transaction_amount = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td.text-right.W90').text    
    if int(transaction_amount.replace(',','')) < 1000:
        break
    else:
        market = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td:nth-child(2) > span.f11px.fVerdana.marginT5').text
        code = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td:nth-child(2) > span.copySelection').text
        title = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td.text-left > a').text
        price = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td.text-right.W80').text
        arrow_status = unit.select('#divList > div.table-responsive.board-table > table > tbody > tr > td')
        
        if unit.find('i', attrs={'class':'fa fa-caret-down'}):
            arrow = '(down)'
        elif unit.find('i', attrs={'class':'fa fa-caret-up'}):
            arrow = '(up)'
        
        price_ratio = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td.text-right.W70').text
        market_cap = unit.select_one('#divList > div.table-responsive.board-table > table > tbody > tr > td:nth-child(8)').text
        sheet.cell(row, 1, market)
        sheet.cell(row, 2, code)
        sheet.cell(row, 3, title)
        sheet.cell(row, 4, price)
        sheet.cell(row, 5, arrow + ' ' + price_ratio)
        sheet.cell(row, 6, transaction_amount)
        sheet.cell(row, 7, market_cap)
        row += 1
        
wb.save(f'D:/lastBox/당일 거래대금 1000억 이상/당일 거래대금 1000억 이상 xlsx/거래대금 1000억 이상_{date_time}.xlsx')

