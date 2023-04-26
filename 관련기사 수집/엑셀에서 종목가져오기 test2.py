import openpyxl as xl

wb = xl.load_workbook('D:/lastBox/당일 거래대금 1000억 이상/당일 거래대금 1000억 이상 xlsx/거래대금 1000억 이상_2023-04-26 14-39-$S.xlsx')

for sheet_nm in wb.sheetnames:
    print('*' * 100)
    print('시트명: ', sheet_nm)
    sheet = wb[sheet_nm]

    for row_data in sheet.iter_rows(min_row=1):
        for cell in row_data:
            print('[' + cell.value + ']')
        print('='*100)
wb.close()